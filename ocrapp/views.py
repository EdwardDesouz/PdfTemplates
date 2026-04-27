import re
import pdfplumber
from django.http import HttpResponse
from django.shortcuts import render
from django.core.files.storage import FileSystemStorage
from openpyxl import Workbook
import os

#Comma and Fullstop for Celine
def swap_separators(value):
    if not value:
        return value
    return value.replace(".", "TEMP").replace(",", ".").replace("TEMP", ",")

#HSCode Extract for Celine
def extract_all_hs_codes(pdf_path):
    HS_X0 = 574 
    HS_X1 = 614
    hs_codes = []

    with pdfplumber.open(pdf_path) as pdf:
        for page_number, page in enumerate(pdf.pages, start=1):
            words = page.extract_words()

            # Find HS Code header
            hs_header_y = None
            for i, word in enumerate(words):
                text = word['text'].strip().upper()
                if text == "HS" and i + 1 < len(words):
                    next_word = words[i + 1]['text'].strip().upper()
                    if next_word == "CODE":
                        hs_header_y = max(word['bottom'], words[i + 1]['bottom'])
                        break
            if hs_header_y is None:
                continue  # no header on this page

            # Collect HS codes below header within column
            for word in words:
                x0, y0 = word['x0'], word['top']
                text = word['text'].strip()
                if y0 > hs_header_y and HS_X0 <= x0 <= HS_X1 and re.fullmatch(r"\d{6,8}", text):
                    hs_codes.append({
                        "page": page_number,
                        "hs_code": text,
                        "top": y0,
                        "bottom": word['bottom']
                    })

    return hs_codes


#Extract Invoice For Celine
def extract_invoice_numbers(pdf_path):
    # Rectangle coordinates for invoice number area
    INV_X0 = 600
    INV_X1 = 1000
    INV_Y0 = 50
    INV_Y1 = 100

    invoice_numbers = []

    with pdfplumber.open(pdf_path) as pdf:
        for page_number, page in enumerate(pdf.pages, start=1):
            words = page.extract_words()

            # Keep only words inside the rectangle
            rect_words = [w for w in words if INV_X0 <= w['x0'] <= INV_X1 and INV_Y0 <= w['top'] <= INV_Y1]

            if not rect_words:
                continue

            # Sort by top, then left to form lines
            rect_words.sort(key=lambda w: (w['top'], w['x0']))

            # Combine words into a single line
            line_text = " ".join(w['text'] for w in rect_words).strip()

            # Look for "DOCUMENT N° :" and capture what comes after
            match = re.search(r"DOCUMENT\s*N°?\s*:?(.+)", line_text, re.IGNORECASE)
            if match:
                invoice_number = match.group(1).strip()  # this is "1164466 RJ"
                invoice_numbers.append({
                    "page": page_number,
                    "invoice_number": invoice_number,
                    "top": rect_words[0]['top'],
                    "bottom": rect_words[-1]['bottom']
                })

    return invoice_numbers


#remove duplicate for linex
def remove_repeated_phrases(text):
    words = text.split()
    i = 0
    while i < len(words):
        for size in range(len(words)//2, 0, -1):
            if i + 2*size <= len(words):
                first = words[i:i+size]
                second = words[i+size:i+2*size]
                if first == second:
                    del words[i+size:i+2*size]
                    # After deletion, check again from current position
                    i = max(i - size, 0)
                    break
        else:
            i += 1
    return " ".join(words)


#Common DropDown


def extract_pdf_data(request):
    if request.method == "POST" and request.FILES.get("pdf"):
        mode = request.POST.get("mode")
        pdf_file = request.FILES["pdf"]

        # ===============================
        # SAVE PDF TO SERVER
        # ===============================
        fs = FileSystemStorage(location='media/pdfs')
        os.makedirs(fs.location, exist_ok=True)
        saved_pdf_name = fs.save(pdf_file.name, pdf_file)
        saved_pdf_path = fs.path(saved_pdf_name)

        # ===============================
        # EXTRACT FULL TEXT FROM PDF
        # ===============================
        full_text = ""
        with pdfplumber.open(saved_pdf_path) as pdf:
            for page in pdf.pages:
                text = page.extract_text()
                if text:
                    full_text += text + "\n"

        extracted_data = []

        # ===============================
        # HAKKO CORPORATION MODE
        # ===============================
        if mode == "Hakko_Corporation":
            blocks = re.split(r"\*FREE OF CHARGE", full_text, flags=re.IGNORECASE)
            for block in blocks:
                block = block.strip()
                if not block:
                    continue
                # Extract COO
                coo_match = re.search(r"MADE IN\s+([A-Z]+)", block, re.IGNORECASE)
                coo = coo_match.group(1).upper() if coo_match else ""
                # Extract Quantity, Unit, Unit Price, Amount from last table line
                table_line_match = re.findall(
                    r"(\d+(?:\.\d+)?)\s+(\w+)\s*\(\s*([\d,]+\.\d+)\s*\)\s*\(\s*([\d,]+\.\d+)\s*\)",
                    block,
                    re.MULTILINE
                )
                if table_line_match:
                    quantity, unit, unit_price, amount = table_line_match[-1]
                    try:
                        quantity = float(quantity.replace(",", ""))
                    except:
                        quantity = 0.0
                    try:
                        unit_price = float(unit_price.replace(",", ""))
                    except:
                        unit_price = 0.0
                    try:
                        amount = float(amount.replace(",", ""))
                    except:
                        amount = 0.0
                else:
                    quantity = 0.0
                    unit = ""
                    unit_price = 0.0
                    amount = 0.0

                extracted_data.append({
                    "COO": coo,
                    "Quantity": quantity,
                    "Unit": unit,
                    "Unit Price": unit_price,
                    "Amount": amount,
                    "HS Code": "" 
                })

        # ===============================
        # lINUX
        # ===============================
        elif mode == "linux":

            lines = full_text.splitlines()

            current_row = None
            start_processing = False

            for line in lines:
                line = line.strip()
                if not line:
                    continue

                # Start after header row
                if not start_processing:
                    if ("Commodity" in line and "Full Description of" in line and "No. of" in line):
                        start_processing = True
                    continue

                # Stop when Shipping Charge appears
                if "Shipping Charge" in line:
                    break

                # HS Code row
                hs_match = re.match(r"^(\d{4,10})\s+(.*)", line)

                if hs_match:

                    # Save previous row
                    if current_row:
                        current_row["Description"] = remove_repeated_phrases(current_row["Description"])
                        extracted_data.append(current_row)

                    hs_code = hs_match.group(1)
                    rest_of_line = hs_match.group(2).strip()

                    tokens = rest_of_line.split()

                    # Extract numeric values
                    numeric_tokens = [
                        t.replace(",", "")
                        for t in tokens
                        if re.match(r"^\d+(?:,\d+)*(?:\.\d+)?$", t)
                    ]

                    no_of_items = numeric_tokens[0] if len(numeric_tokens) >= 1 else "0"
                    unit_value = numeric_tokens[1] if len(numeric_tokens) >= 2 else "0"
                    total_value = numeric_tokens[2] if len(numeric_tokens) >= 3 else "0"

                    # Extract COO (last token usually country)
                    coo = ""
                    if tokens:
                        possible_coo = tokens[-1]
                        if re.match(r"^[A-Z]{2,}$", possible_coo):
                            coo = possible_coo

                    # Description before numbers
                    first_num_index = next(
                        (i for i, t in enumerate(tokens) if re.match(r"^\d+(?:,\d+)*(?:\.\d+)?$", t)),
                        len(tokens)
                    )

                    description_tokens = tokens[:first_num_index]
                    description = " ".join(description_tokens)

                    current_row = {
                        "HS Code": hs_code,
                        "Description": description,
                        "Product Composition": "",
                        "No. of Items": no_of_items,
                        "Unit Value": unit_value,
                        "Total Value": total_value,
                        "COO": coo
                    }

                else:
                    if current_row:
                        current_row["Description"] += " " + line

            # Append last row
            if current_row:
                current_row["Description"] = remove_repeated_phrases(current_row["Description"])
                extracted_data.append(current_row)
    
        # ===============================
        # THOSIBA IMPORT
        # ===============================
        elif mode == "Toshiba_Import":
            extracted_data = []

            with pdfplumber.open(saved_pdf_path) as pdf:
                full_text = ""
                for page in pdf.pages:
                    text = page.extract_text()
                    if text:
                        full_text += text + "\n"

            # Split based on PO numbers
            parts = re.split(r"P/O\s*NO:\s*(\d+)", full_text)

            # Quantity + Unit pattern
            unit_pattern = r"\b(\d{1,3}(?:,\d{3})*|\d+)\s*(PCS|NOS|UNITS|BOX|PACK|SET|PAIR|KG|G|TON|L|ML|M|CM|FT)\b"

            for i in range(1, len(parts), 2):
                po_number = parts[i]
                text_after = parts[i + 1]

                lines = text_after.strip().splitlines()

                description = ""
                quantity = ""
                unit = ""
                unit_price = ""
                amount = ""
                diffused_in = ""
                assembled_in = ""

                for idx, line in enumerate(lines):
                    clean_line = line.strip()

                    if not clean_line:
                        continue

                    # ❌ Skip unwanted lines
                    if re.search(r"OUR\s+REF\s+NO", clean_line, re.IGNORECASE):
                        continue

                    if "@" in clean_line:
                        continue

                    # ✅ Description
                    if not description:
                        description = re.sub(
                            r"\s+\d+(\.\d+)?(\s+\d+(\.\d+)?)*$",
                            "",
                            clean_line
                        )

                    # ✅ Quantity + Unit
                    qty_match = re.search(unit_pattern, clean_line, re.IGNORECASE)

                    if qty_match:
                        quantity = qty_match.group(1)   # keep comma
                        unit = qty_match.group(2).upper()

                        # ✅ Extract Unit Price & Amount
                        after_qty = clean_line[qty_match.end():].strip()

                        numbers = re.findall(
                            r"\d{1,3}(?:,\d{3})*\.\d+|\d+\.\d+",
                            after_qty
                        )

                        if len(numbers) == 1:
                            amount = numbers[0]   # single value → treat as amount
                        elif len(numbers) >= 2:
                            unit_price = numbers[0]
                            amount = numbers[1]

                        # ============================
                        # 🔥 FINAL DIFFUSED / ASSEMBLED FIX
                        # ============================

                        # Look ahead more lines for safety
                        next_lines = lines[idx+1: idx+8]

                        combined_text = " ".join(
                            [l.strip() for l in next_lines if l.strip()]
                        ).upper()

                        # 🔥 Normalize broken OCR text
                        combined_text = re.sub(
                            r'(?<=\b[A-Z])\s+(?=[A-Z]\b)',
                            '',
                            combined_text
                        )

                        # ✅ Step 1: Extract ALL "IN:" values
                        in_values = re.findall(r"IN[:\s]*([A-Z]+)", combined_text)

                        if len(in_values) >= 1:
                            diffused_in = in_values[0]

                        if len(in_values) >= 2:
                            assembled_in = in_values[1]

                        # ✅ Step 2: Override using clean "ASSEMBLED IN" if available
                        assm_match = re.search(
                            r"ASSEMBLED\s*IN[:\s]*([A-Z]+)",
                            combined_text
                        )

                        if assm_match:
                            assembled_in = assm_match.group(1)

                        break  # stop after finding quantity row

                extracted_data.append({
                    "PO Number": po_number,
                    "Description": description,
                    "Quantity": quantity,
                    "Unit": unit,
                    "Unit Price": unit_price,
                    "Amount": amount,
                    "Diffused In": diffused_in,
                    "Assembled In": assembled_in
                })
            
        # ===============================
        # CELINE MODE
        # ===============================
 
        elif mode == "Celine":
            hs_codes_positions = extract_all_hs_codes(saved_pdf_path)  # your existing function
            invoice_numbers_positions = extract_invoice_numbers(saved_pdf_path)
            # print("invoice_numbers_positions:",invoice_numbers_positions)
            lines = full_text.splitlines()

            seen_invoice_numbers = set() 
            hs_index = 0  # pointer to assign HS codes one by one
            inv_index = 0
            for line in lines:
                line = line.strip()
                if not line:
                    continue

                # Extract COO
                coo_match = re.search(r"Made in\s+([A-Z]{2})", line, re.IGNORECASE)
                if not coo_match:
                    continue

                coo = coo_match.group(1).upper()

                # Remove "Made in XX"
                remaining_text = re.sub(
                    r"Made in\s+[A-Z]{2}", "", line, flags=re.IGNORECASE
                ).strip()

                parts = remaining_text.split()

                quantity = parts[0] if len(parts) >= 1 else ""
                unit = parts[1] if len(parts) >= 2 else ""
                unit_price = parts[2] if len(parts) >= 3 else ""
                #print("unit_price",unit_price)
                unit_price_1 = swap_separators(unit_price)
                #print("unit_price_1",unit_price_1)
                amount = parts[3] if len(parts) >= 4 else ""

                # Assign HS Code in order
                hs_code = ""
                if hs_index < len(hs_codes_positions):
                    hs_code = hs_codes_positions[hs_index]['hs_code']
                    hs_index += 1  # move to next HS code for the next line

                # invoice_number = ""
                # if inv_index < len(invoice_numbers_positions):
                #     invoice_number = invoice_numbers_positions[inv_index]['invoice_number']
                #     inv_index += 1
                #     print("invoice_number:",invoice_number)
                invoice_number = ""
                while inv_index < len(invoice_numbers_positions):
                    potential_invoice = invoice_numbers_positions[inv_index]['invoice_number']
                    inv_index += 1
                    if potential_invoice not in seen_invoice_numbers:
                        invoice_number = potential_invoice
                        seen_invoice_numbers.add(invoice_number)
                        break

                extracted_data.append({
                    "COO": coo,
                    "Quantity": quantity,
                    "Unit": unit,
                    "Unit Price": "",
                    "Amount": unit_price_1, 
                    "HS Code": hs_code,
                    "Invoice Number": invoice_number 

                })

                
        # ===============================
        # NNRGLOBAL MODE
        # ===============================


        elif mode == "Nnr_Global_Panasonic":
            pattern = r"([^\r\n]+)\s*[\r\n]+\s*Made In\s+([A-Za-z]+)"
            matches = re.findall(pattern, full_text)

            for index, (line_data, country) in enumerate(matches, start=1):
                # Remove leading numbers/special characters until first letter
                clean_line_match = re.search(r"[A-Za-z].*", line_data)
                clean_line_text = clean_line_match.group() if clean_line_match else line_data.strip()

                # Initialize variables
                description = ""
                quantity = ""
                unit = ""
                amount = ""
                rest = []

                # Split description and numeric parts
                first_number_match = re.search(r"\d", clean_line_text)
                if first_number_match:
                    num_index = first_number_match.start()
                    description = clean_line_text[:num_index].strip()
                    rest = clean_line_text[num_index:].split()
                    quantity = rest[0] if len(rest) > 0 else ""
                    unit = rest[1] if len(rest) > 1 else ""
                    amount = rest[2] if len(rest) > 2 else ""
                else:
                    description = clean_line_text.strip()

                # Append to extracted_data
                extracted_data.append({
                    "COO": country.upper(),
                    "Description": description,
                    "Quantity": quantity,
                    "Unit": unit,
                    "Amount": amount
                })
        
        # ===============================
        # MARIENTRANS MODE
        # ===============================
        # elif mode == "Marinetrans":
        #     extracted_data = []

        #     # Regex to find "HS Code" header in table
        #     header_pattern = re.compile(r"\bHS\s*Code\b", re.IGNORECASE)
            
        #     # Regex to match actual HS Codes: 6 digits + space + 2 digits
        #     hscode_pattern = re.compile(r"^\d{6}\s\d{2}$")
        #     #hscode_pattern = re.compile(r"^\d{6}\s\d{2}")

        #     with pdfplumber.open(saved_pdf_path) as pdf:
        #         for page in pdf.pages:
        #             tables = page.extract_tables()
        #             if not tables:
        #                 continue

        #             for table in tables:
        #                 # Check if table has HS Code header
        #                 table_text = "\n".join(
        #                     [" ".join([cell if cell else "" for cell in row]) for row in table]
        #                 )

        #                 if header_pattern.search(table_text):
        #                     # Extract rows with HS Code
        #                     for row in table:
        #                         for idx, cell in enumerate(row):
        #                             if cell:
        #                                 cell_text = cell.strip()
        #                                 if hscode_pattern.match(cell_text):
        #                                     # Take all data after HS Code
        #                                     # row_after_hscode = [c if c else "" for c in row[idx+1:]]
        #                                     parts = cell_text.split()

        #                             # ✅ Always extract proper HS Code
        #                                     hs_code = " ".join(parts[:2])

        #                             # ✅ Case 1: merged row (all data in same cell)
        #                                     if len(parts) > 2:
        #                                         remaining = parts[2:]
        #                                     else:
        #                                 # ✅ Case 2: normal table structure
        #                                         remaining = [c if c else "" for c in row[idx+1:]]

        #                                     # Map row data to proper columns using indices
        #                                     # row_dict = {
        #                                     #     "HSCode": cell_text,
        #                                     #     "SG Post": row_after_hscode[0] if len(row_after_hscode) > 0 else "",
        #                                     #     "CoO": row_after_hscode[1] if len(row_after_hscode) > 1 else "",
        #                                     #     "Batch Number": row_after_hscode[2] if len(row_after_hscode) > 2 else "",
        #                                     #     "Batch Number": row_after_hscode[3] if len(row_after_hscode) > 3 else "",
        #                                     #     "UOM": row_after_hscode[4] if len(row_after_hscode) > 4 else "",
        #                                     #     "Weight": row_after_hscode[5] if len(row_after_hscode) > 5 else "",
        #                                     #     "W.UOM": row_after_hscode[6] if len(row_after_hscode) > 6 else "",
        #                                     #     "Unit Price": row_after_hscode[7] if len(row_after_hscode) > 7 else "",
        #                                     #     "Currency": row_after_hscode[8] if len(row_after_hscode) > 8 else "",
        #                                     #     "Customs Value": row_after_hscode[9] if len(row_after_hscode) > 9 else ""
        #                                     # }
        #                                     row_dict = {
        #                                 "HSCode": hs_code,
        #                                 "SG Post": remaining[0] if len(remaining) > 0 else "",
        #                                 "CoO": remaining[1] if len(remaining) > 1 else "",
        #                                 "Batch Number": remaining[2] if len(remaining) > 2 else "",
        #                                 "Batch Number 2": remaining[3] if len(remaining) > 3 else "",
        #                                 "UOM": remaining[4] if len(remaining) > 4 else "",
        #                                 "Weight": remaining[5] if len(remaining) > 5 else "",
        #                                 "W.UOM": remaining[6] if len(remaining) > 6 else "",
        #                                 "Unit Price": remaining[7] if len(remaining) > 7 else "",
        #                                 "Currency": remaining[8] if len(remaining) > 8 else "",
        #                                 "Customs Value": remaining[9] if len(remaining) > 9 else ""
        #                             }


        #                                     extracted_data.append(row_dict)
        #                                     break  # Only first HS Code per row

        elif mode == "Marinetrans":
            extracted_data = []

    # Only HS Code detection
            hscode_pattern = re.compile(r"^\d{6}\s\d{2}")

            with pdfplumber.open(saved_pdf_path) as pdf:
                for page in pdf.pages:
                    tables = page.extract_tables()
                    if not tables:
                        continue

                    for table in tables:
                        for row in table:
                            for idx, cell in enumerate(row):
                                if cell:
                                    cell_text = cell.strip()

                                    # ✅ Detect HS Code directly
                                    if hscode_pattern.match(cell_text):

                                        parts = cell_text.split()
                                        hs_code = " ".join(parts[:2])

                                        # ✅ Handle merged + normal rows
                                        if len(parts) > 2:
                                            remaining = parts[2:]
                                        else:
                                            remaining = [c if c else "" for c in row[idx+1:]]

                                        row_dict = {
                                            "HSCode": hs_code,
                                            "SG Post": remaining[0] if len(remaining) > 0 else "",
                                            "CoO": remaining[1] if len(remaining) > 1 else "",
                                            "Batch Number": remaining[2] if len(remaining) > 2 else "",
                                            "Batch Number 2": remaining[3] if len(remaining) > 3 else "",
                                            "UOM": remaining[4] if len(remaining) > 4 else "",
                                            "Weight": remaining[5] if len(remaining) > 5 else "",
                                            "W.UOM": remaining[6] if len(remaining) > 6 else "",
                                            "Unit Price": remaining[7] if len(remaining) > 7 else "",
                                            "Currency": remaining[8] if len(remaining) > 8 else "",
                                            "Customs Value": remaining[9] if len(remaining) > 9 else ""
                                        }

                                        extracted_data.append(row_dict)
                                        break  # Only first HS Code per row

        # ===============================
        # TOSHIBA MODE
        # ===============================
        elif mode == "Nnr_Toshiba_1":
            # Extract Customer PO Data
            # po_data = []
            # with pdfplumber.open(saved_pdf_path) as pdf:
            #     for page in pdf.pages:
            #         text = page.extract_text()
            #         if not text:
            #             continue
            #         lines = text.splitlines()
            #         capture = False
            #         for i in range(len(lines)):
            #             line = lines[i].strip()
            #             if not line:
            #                 continue
            #             if not capture and re.search(r"Item\s+Customer'?s?\s+P/O\s+No\.?", line, re.IGNORECASE):
            #                 capture = True
            #                 continue
            #             if capture:
            #                 po_match = re.match(r"^\d+\s+([A-Z]+[A-Z0-9-]+)\s+(.*)", line)
            #                 if po_match:
            #                     po_number = po_match.group(1)
            #                     main_row = po_match.group(2).split()
            #                     quantity_index = None
            #                     for idx, value in enumerate(main_row):
            #                         if re.match(r"^\d{1,3}(,\d{3})*$", value):
            #                             quantity_index = idx
            #                             break
            #                     if quantity_index is not None and len(main_row) >= quantity_index + 3:
            #                         description = " ".join(main_row[:quantity_index])
            #                         quantity = main_row[quantity_index]
            #                         unit_price = main_row[quantity_index + 1]
            #                         raw_amount = main_row[quantity_index + 2]
            #                         amount = re.sub(r"[^\d.,]", "", raw_amount)
            #                         customer_pn = ""
            #                         toshiba_pn = ""
            #                         if i + 1 < len(lines):
            #                             parts = lines[i + 1].strip().split()
            #                             if len(parts) >= 1:
            #                                 customer_pn = parts[0]
            #                             if len(parts) >= 2:
            #                                 toshiba_pn = parts[1].replace("RoHS", "")
            #                         po_data.append({
            #                             "PO": po_number,
            #                             "Description": description,
            #                             "Quantity": quantity,
            #                             "UnitPrice": unit_price,
            #                             "Amount": amount,
            #                             "CustomerPN": customer_pn,
            #                             "ToshibaPN": toshiba_pn
            #                         })
            # # print("po_data:",po_data)
                # Extract Customer PO Data
            # po_data = []
            # with pdfplumber.open(saved_pdf_path) as pdf:
            #     for page in pdf.pages:
            #         text = page.extract_text()
            #         if not text:
            #             continue
            #         lines = text.splitlines()
            #         capture = False
            #         for i in range(len(lines)):
            #             line = lines[i].strip()
            #             if not line:
            #                 continue
            #             if not capture and re.search(r"Item\s+Customer'?s?\s+P/O\s+No\.?", line, re.IGNORECASE):
            #                 capture = True
            #                 continue
            #             if capture:
            #                 #po_match = re.match(r"^\d+\s+([A-Z]+[A-Z0-9-]+)\s+(.*)", line)
            #                 po_match = re.match(r"^\d+\s+([A-Z0-9-]+)\s+(.+)$", line)
            #                 if po_match:
            #                     po_number = po_match.group(1)
            #                     rest = po_match.group(2)
            #                     main_row = po_match.group(2).split()
            #                     quantity_index = None
            #                     for idx, value in enumerate(main_row):
            #                         if re.match(r"^\d{1,3}(?:,\d{3})*(?:\.\d+)?$", value):
            #                             quantity_index = idx
            #                             break
            #                     if quantity_index is not None and len(main_row) >= quantity_index + 3:
            #                         description = " ".join(main_row[:quantity_index])
            #                         quantity = main_row[quantity_index]
            #                         unit_price = main_row[quantity_index + 1]
            #                         raw_amount = main_row[quantity_index + 2]
            #                         amount = re.sub(r"[^\d.,]", "", raw_amount)
            #                         customer_pn = ""
            #                         toshiba_pn = ""
            #                         if i + 1 < len(lines):
            #                             parts = lines[i + 1].strip().split()
            #                             if len(parts) >= 1:
            #                                 customer_pn = parts[0]
            #                             if len(parts) >= 2:
            #                                 toshiba_pn = parts[1].replace("RoHS", "")
            #                         po_data.append({
            #                             "PO": po_number,
            #                             "Description": description,
            #                             "Quantity": quantity,
            #                             "UnitPrice": unit_price,
            #                             "Amount": amount,
            #                             "CustomerPN": customer_pn,
            #                             "ToshibaPN": toshiba_pn
            #                         })

            # # ===============================
            # # AGGREGATE BY CUSTOMERPN
            # # ===============================
            # aggregated_po_data = {}

            # for item in po_data:
            #     cust_pn = item["CustomerPN"]
            #     if not cust_pn:
            #         continue  # skip empty CustomerPN
            #     # Convert Quantity and Amount to float safely
            #     try:
            #         quantity = float(item["Quantity"].replace(",", ""))
            #     except:
            #         quantity = 0.0
            #     try:
            #         amount = float(item["Amount"].replace(",", ""))
            #     except:
            #         amount = 0.0
            #     try:
            #         unit_price = float(item["UnitPrice"].replace(",", ""))
            #     except:
            #         unit_price = 0.0

            #     if cust_pn in aggregated_po_data:
            #         # Sum quantities and amounts
            #         aggregated_po_data[cust_pn]["Quantity"] += quantity
            #         aggregated_po_data[cust_pn]["Amount"] += amount
            #         # Merge POs
            #         aggregated_po_data[cust_pn]["PO"] += f"/{item['PO']}"
            #     else:
            #         aggregated_po_data[cust_pn] = {
            #             "PO": item["PO"],
            #             "Description": item["Description"],
            #             "Quantity": quantity,
            #             "UnitPrice": unit_price,
            #             "Amount": amount,
            #             "CustomerPN": cust_pn,
            #             "ToshibaPN": item["ToshibaPN"]
            #         }

            # # Convert aggregated dict to list for Excel
            # po_data = list(aggregated_po_data.values())


            po_data = []

            with pdfplumber.open(saved_pdf_path) as pdf:
                for page in pdf.pages:
                    text = page.extract_text()
                    if not text:
                        continue
                    lines = [line.strip() for line in text.splitlines() if line.strip()]
                    capture = False
                    i = 0
                    while i < len(lines):
                        line = lines[i]

                        # Detect start of PO table
                        if not capture and re.search(r"Item\s+Customer'?s?\s+P/O\s+No\.?", line, re.IGNORECASE):
                            capture = True
                            i += 1
                            continue

                        if capture:
                            # Match line starting with number + PO number
                            po_match = re.match(r"^\d+\s+([A-Z0-9-]+)\s+(.+)$", line)
                            if po_match:
                                po_number = po_match.group(1)
                                rest = po_match.group(2)

                                # Split from the right to reliably extract Quantity, UnitPrice, Amount
                                try:
                                    description_part, quantity, unit_price, raw_amount = rest.rsplit(maxsplit=3)
                                    description = description_part
                                    amount = re.sub(r"[^\d.,]", "", raw_amount)  # remove asterisks or symbols
                                except ValueError:
                                    # fallback if line structure is irregular
                                    description = rest
                                    quantity = unit_price = amount = "0"

                                # Next line contains Customer PN and Toshiba PN (if exists)
                                customer_pn = ""
                                toshiba_pn = ""
                                if i + 1 < len(lines):
                                    next_line = lines[i + 1].strip()
                                    parts = next_line.split()
                                    if len(parts) >= 2:
                                        customer_pn = parts[0]
                                        toshiba_pn = parts[1].replace("RoHS", "")
                                        i += 1  # skip this line as it's processed
                                    elif len(parts) == 1:
                                        customer_pn = parts[0]
                                        i += 1

                                po_data.append({
                                    "PO": po_number,
                                    "Description": description,
                                    "Quantity": quantity,
                                    "UnitPrice": unit_price,
                                    "Amount": amount,
                                    "CustomerPN": customer_pn,
                                    "ToshibaPN": toshiba_pn
                                })

                        i += 1

            # ===============================
            # AGGREGATE BY CUSTOMERPN
            # ===============================
            aggregated_po_data = {}

            for item in po_data:
                cust_pn = item["CustomerPN"]
                if not cust_pn:
                    continue  # skip empty CustomerPN

                # Safely convert numbers
                try:
                    quantity = float(item["Quantity"].replace(",", ""))
                except:
                    quantity = 0.0
                try:
                    amount = float(item["Amount"].replace(",", ""))
                except:
                    amount = 0.0
                try:
                    unit_price = float(item["UnitPrice"].replace(",", ""))
                except:
                    unit_price = 0.0

                if cust_pn in aggregated_po_data:
                    aggregated_po_data[cust_pn]["Quantity"] += quantity
                    aggregated_po_data[cust_pn]["Amount"] += amount
                    aggregated_po_data[cust_pn]["PO"] += f"/{item['PO']}"
                else:
                    aggregated_po_data[cust_pn] = {
                        "PO": item["PO"],
                        "Description": item["Description"],
                        "Quantity": quantity,
                        "UnitPrice": unit_price,
                        "Amount": amount,
                        "CustomerPN": cust_pn,
                        "ToshibaPN": item["ToshibaPN"]
                    }

            # Convert to list for Excel or further processing
            po_data = list(aggregated_po_data.values())

            # # Extract Package Data

#old working Code 1

            package_data = []
            with pdfplumber.open(saved_pdf_path) as pdf:
                for page in pdf.pages:
                    text = page.extract_text()
                    if not text:
                        continue
                    lines = [line.strip() for line in text.splitlines() if line.strip()]
                    capture = False
                    i = 0
                    while i < len(lines):
                        line = lines[i]
                        # if not capture and "Package" in line and "Customer" in line and "Part" in line:
                        if not capture and re.search(r"Package\s+Customer\s+Part", line, re.IGNORECASE):
                            capture = True
                            i += 1
                            continue
                        if capture:
                            package_match = re.match(r"^(\d+)\s*(.*)", line)
                            if package_match:
                                package_number = package_match.group(1)
                                j = i + 1
                                package_lines = []
                                
                                while j < len(lines):
                                    next_line = lines[j]
                                    if re.match(r"^\d+\s*", next_line):
                                        break
                                 # Skip irrelevant lines
                                    if re.search(r"TOTAL PACKAGES|PAGE\s*\d+", next_line, re.IGNORECASE):
                                        j += 1
                                        continue
                                    package_lines.append(next_line)
                                    j += 1
                                customer_part = package_lines[0] if len(package_lines) > 0 else ""
                                coo = package_lines[1] if len(package_lines) > 1 else ""
                                if customer_part.strip() or coo.strip():
                                    part_pattern = re.match(
                                        r"(\S+)\s+(.+)\s+(\d+)\s+([\d\.]+)\s+([\d\.]+)",customer_part
                                    )
                                    if part_pattern:
                                        part_number = part_pattern.group(1)
                                        description = part_pattern.group(2)
                                        quantity = part_pattern.group(3)
                                        weight = part_pattern.group(4)
                                        unit_price = part_pattern.group(5)
                                    else:
                                        parts = customer_part.split()
                                        part_number = parts[0] if len(parts) > 0 else ""
                                        description = " ".join(parts[1:-3]) if len(parts) > 4 else ""
                                        quantity = parts[-3] if len(parts) > 2 else ""
                                        weight = parts[-2] if len(parts) > 1 else ""
                                        unit_price = parts[-1] if len(parts) > 0 else ""
                                    coo_clean = coo.upper().replace("MADE IN ", "").strip()
                                    package_data.append({
                                        "PartNumber": part_number,
                                        "Description": description,
                                        "Quantity": quantity,
                                        "Weight": weight,
                                        "UnitPrice": unit_price,
                                        "Coo": coo_clean
                                    })
                                i = j
                                continue
                        i += 1


#new code with some issues 2

        # package_data = []
        # with pdfplumber.open(saved_pdf_path) as pdf:
        #     for page in pdf.pages:
        #         text = page.extract_text()
        #         if not text:
        #             continue
        #         lines = [line.strip() for line in text.splitlines() if line.strip()]
        #         capture = False
        #         i = 0
        #         while i < len(lines):
        #             line = lines[i]
        #             # Detect start of package table
        # #             if not capture and re.search(r"Package\s+Customer\s+Part", line, re.IGNORECASE):
        #             if not capture and "Package" in line and "Customer" in line and "Part" in line:
        #                 capture = True
        #                 i += 1
        #                 continue

        #             if capture:
        #                 # Detect a new package by numeric line
        #                 #package_match = re.match(r"^(\d+)\s*(.*)$", line)
        #                 package_match = re.match(r"^(\d+(?:-\d+)?)\s*(.*)$", line)
        #                 if package_match:
        #                     package_number = package_match.group(1)
        #                     j = i + 1
        #                     package_lines = []


        #                     # Collect all lines until next package or end
        #                     while j < len(lines):
        #                         next_line = lines[j]
        #                         if re.match(r"^\d+\s*$", next_line):  # next package
        #                             break
        #                         # Skip irrelevant lines
        #                         if re.search(r"TOTAL PACKAGES|PAGE\s*\d+", next_line, re.IGNORECASE):
        #                             j += 1
        #                             continue
        #                         package_lines.append(next_line)
        #                         print("package_lines:",package_lines)
        #                         j += 1
        #                     customer_part = package_lines[0] if len(package_lines) > 0 else ""
        #                     coo = package_lines[1] if len(package_lines) > 1 else ""
        #                     if customer_part.strip() or coo.strip():
        #                         part_pattern = re.match(
        #                             r"(\S+)\s+(.+)\s+(\d+)\s+([\d\.]+)\s+([\d\.]+)",customer_part)
        #                         if part_pattern:
        #                             part_number = part_pattern.group(1)
        #                             description = part_pattern.group(2)
        #                             quantity = part_pattern.group(3)
        #                             weight = part_pattern.group(4)
        #                             unit_price = part_pattern.group(5)
        #                         else:
        #                             parts = customer_part.split()
        #                             part_number = parts[0] if len(parts) > 0 else ""
        #                             description = " ".join(parts[1:-3]) if len(parts) > 4 else ""
        #                             quantity = parts[-3] if len(parts) > 2 else ""
        #                             weight = parts[-2] if len(parts) > 1 else ""
        #                             unit_price = parts[-1] if len(parts) > 0 else ""
        #                         coo_clean = coo.upper().replace("MADE IN ", "").strip()
        #                         package_data.append({
        #                             "PartNumber":part_number ,
        #                             "Description":description,
        #                             "Quantity": quantity,
        #                             "Weight": weight,
        #                             "UnitPrice":unit_price ,
        #                             "Coo":coo_clean,

        #                         })

        #                     i = j
        #                     continue
        #             i += 1

#old and new code incorrparate 3


        # package_data = []

        # with pdfplumber.open(saved_pdf_path) as pdf:
        #     for page in pdf.pages:
        #         text = page.extract_text()
        #         if not text:
        #             continue

        #         # Clean lines, remove empty lines
        #         lines = [line.strip() for line in text.splitlines() if line.strip()]

        #         capture = False
        #         i = 0
        #         while i < len(lines):
        #             line = lines[i]

        #             # Detect start of package table
        #             if not capture and "Package" in line and "Customer" in line and "Part" in line:
        #                 capture = True
        #                 i += 1
        #                 continue

        #             if capture:
        #                 # Detect a new package by numeric line (stricter regex)
                         ## package_match = re.match(r"^(\d+)\s*(.*)$", line)
        #                 if package_match:
        #                     package_number = package_match.group(1)
        #                     j = i + 1
        #                     package_lines = []

        #                     # Collect all lines until next package or end
        #                     while j < len(lines):
        #                         next_line = lines[j]

        #                         # Stop if next package line
        #                         if re.match(r"^\d+\s*$", next_line):
        #                             break

        #                         # Skip irrelevant lines
        #                         if re.search(r"TOTAL PACKAGES|PAGE\s*\d+", next_line, re.IGNORECASE):
        #                             j += 1
        #                             continue

        #                         package_lines.append(next_line)
        #                         j += 1

        #                     # Extract customer part and COO
        #                     customer_part = package_lines[0] if len(package_lines) > 0 else ""
        #                     coo = package_lines[1] if len(package_lines) > 1 else ""

        #                     if customer_part.strip() or coo.strip():
        #                         # Try strict pattern first
        #                         part_pattern = re.match(
        #                             r"(\S+)\s+(.+)\s+(\d+)\s+([\d\.]+)\s+([\d\.]+)",
        #                             customer_part
        #                         )
        #                         if part_pattern:
        #                             part_number = part_pattern.group(1)
        #                             description = part_pattern.group(2)
        #                             quantity = part_pattern.group(3)
        #                             weight = part_pattern.group(4)
        #                             unit_price = part_pattern.group(5)
        #                         else:
        #                             # Fallback split method
        #                             parts = customer_part.split()
        #                             part_number = parts[0] if len(parts) > 0 else ""
        #                             description = " ".join(parts[1:-3]) if len(parts) > 4 else ""
        #                             quantity = parts[-3] if len(parts) > 2 else ""
        #                             weight = parts[-2] if len(parts) > 1 else ""
        #                             unit_price = parts[-1] if len(parts) > 0 else ""

        #                         coo_clean = coo.upper().replace("MADE IN ", "").strip()

        #                         package_data.append({
        #                             "PartNumber": part_number,
        #                             "Description": description,
        #                             "Quantity": quantity,
        #                             "Weight": weight,
        #                             "UnitPrice": unit_price,
        #                             "Coo": coo_clean
        #                         })

        #                     i = j
        #                     continue

        #             i += 1

#4

        # package_data = []

        # with pdfplumber.open(saved_pdf_path) as pdf:
        #     capture = False

        #     for page in pdf.pages:
        #         text = page.extract_text()
        #         if not text:
        #             continue

        #         lines = [line.strip() for line in text.splitlines() if line.strip()]
        #         i = 0

        #         while i < len(lines):
        #             line = lines[i]

        #             # Detect start of package table
        #             if not capture and re.search(r"Package\s+Customer\s+Part", line, re.IGNORECASE):
        #                 capture = True
        #                 i += 1
        #                 continue

        #             if capture:
        #                 # Detect package number or range at start of line
        #                 package_match = re.match(r"^(\d+(?:-\d+)?)\b", line)

        #                 # Prevent quantity-only lines (like 7500) from being treated as package
        #                 if package_match and not (line.isdigit() and int(line) > 1000):
        #                     raw_package_number = package_match.group(1)

        #                     # Expand range if needed
        #                     if "-" in raw_package_number:
        #                         start, end = map(int, raw_package_number.split("-"))
        #                         package_numbers = list(range(start, end + 1))
        #                     else:
        #                         package_numbers = [int(raw_package_number)]

        #                     j = i + 1
        #                     package_lines = []

        #                     # Collect lines until next package
        #                     while j < len(lines):
        #                         next_line = lines[j]

        #                         # Stop at next package number
        #                         if re.match(r"^\d+(?:-\d+)?\b", next_line) and not (
        #                             next_line.isdigit() and int(next_line) > 1000
        #                         ):
        #                             break

        #                         # Skip totals and page headers
        #                         if re.search(r"TOTAL PACKAGES|PAGE\s*\d+", next_line, re.IGNORECASE):
        #                             j += 1
        #                             continue

        #                         package_lines.append(next_line)
        #                         j += 1

        #                     # Extract relevant data from collected lines
        #                     customer_part = package_lines[0] if len(package_lines) > 0 else ""
        #                     coo = package_lines[1] if len(package_lines) > 1 else ""

        #                     if customer_part.strip() or coo.strip():
        #                         # Try strict pattern first
        #                         part_pattern = re.match(
        #                             r"(\S+)\s+(.+)\s+(\d+)\s+([\d\.]+)\s+([\d\.]+)",
        #                             customer_part
        #                         )
        #                         if part_pattern:
        #                             part_number = part_pattern.group(1)
        #                             description = part_pattern.group(2)
        #                             quantity = part_pattern.group(3)
        #                             weight = part_pattern.group(4)
        #                             unit_price = part_pattern.group(5)
        #                         else:
        #                             # Fallback split method
        #                             parts = customer_part.split()
        #                             part_number = parts[0] if len(parts) > 0 else ""
        #                             description = " ".join(parts[1:-3]) if len(parts) > 4 else ""
        #                             quantity = parts[-3] if len(parts) > 2 else ""
        #                             weight = parts[-2] if len(parts) > 1 else ""
        #                             unit_price = parts[-1] if len(parts) > 0 else ""

        #                         coo_clean = coo.upper().replace("MADE IN ", "").strip()

        #                         package_data.append({
        #                             "PartNumber": part_number,
        #                             "Description": description,
        #                             "Quantity": quantity,
        #                             "Weight": weight,
        #                             "UnitPrice": unit_price,
        #                             "Coo": coo_clean
        #                         })

        #                     i = j
        #                     continue

        #             i += 1

#5
        # package_data = []
        # with pdfplumber.open(saved_pdf_path) as pdf:
        #     for page in pdf.pages:
        #         text = page.extract_text()
        #         if not text:
        #             continue

        #         # Clean and split lines
        #         lines = [line.strip() for line in text.splitlines() if line.strip()]
        #         capture = False
        #         i = 0
        #         while i < len(lines):
        #             line = lines[i]

        #             # Start capturing after the table header
        #             if not capture and "Package" in line and "Customer" in line and "Part" in line:
        #                 capture = True
        #                 i += 1
        #                 continue

        #             if capture:
        #                 # Detect start of a new package
        #                 package_match = re.match(r"^(\d+(?:-\d+)?)\s*(.*)$", line)
        #                 if package_match:
        #                     j = i + 1
        #                     package_lines = []

        #                     # Collect all lines until next numeric line or end
        #                     while j < len(lines):
        #                         next_line = lines[j]

        #                         # Stop if next package starts
        #                         if re.match(r"^\d+\s*$", next_line):
        #                             break

        #                         # Skip irrelevant lines
        #                         if re.search(r"TOTAL PACKAGES|PAGE\s*\d+|SHIPPING MARKS|INV\.", next_line, re.IGNORECASE):
        #                             j += 1
        #                             continue

        #                         package_lines.append(next_line)
        #                         j += 1

        #                     # Parse package_lines
        #                     current_coo = ""
        #                     for pline in package_lines:
        #                         pline = pline.strip()
        #                         if not pline:
        #                             continue

        #                         # COO line
        #                         if pline.upper().startswith("MADE IN"):
        #                             current_coo = pline.upper().replace("MADE IN", "").strip()
        #                             continue

        #                         # Part line: PartNumber Description Quantity Weight UnitPrice
        #                         part_match = re.match(r"(\S+)\s+(.+?)\s+(\d+)\s+([\d\.]+)\s+([\d\.]+)", pline)
        #                         if part_match:
        #                             part_number = part_match.group(1)
        #                             print("part_number:",part_number)
        #                             description = part_match.group(2)
        #                             print("description:",description)
        #                             quantity = part_match.group(3)
        #                             print("quantity:",quantity)
        #                             weight = part_match.group(4)
        #                             print("weight:",weight)
        #                             unit_price = part_match.group(5)
        #                             print("unit_price:",unit_price)

        #                             package_data.append({
        #                                 "PartNumber": part_number,
        #                                 "Description": description,
        #                                 "Quantity": quantity,
        #                                 "Weight": weight,
        #                                 "UnitPrice": unit_price,
        #                                 "Coo": current_coo
        #                             })
        #                         else:
        #                             # Fallback parsing if regex fails
        #                             parts = pline.split()
        #                             part_number = parts[0] if len(parts) > 0 else ""
        #                             print("part_number:",part_number)
        #                             description = " ".join(parts[1:-3]) if len(parts) > 4 else ""
        #                             print("description:",description)
        #                             quantity = parts[-3] if len(parts) > 2 else ""
        #                             print("quantity:",quantity)
        #                             weight = parts[-2] if len(parts) > 1 else ""
        #                             print("weight:",weight)
        #                             unit_price = parts[-1] if len(parts) > 0 else ""
        #                             print("unit_price:",unit_price)
        #                             package_data.append({
        #                                 "PartNumber": part_number,
        #                                 "Description": description,
        #                                 "Quantity": quantity,
        #                                 "Weight": weight,
        #                                 "UnitPrice": unit_price,
        #                                 "Coo": current_coo
        #                             })

        #                     # Move to next package
        #                     i = j
        #                     continue
        #             i += 1


#6
        # package_data = []

        # with pdfplumber.open(saved_pdf_path) as pdf:
        #     for page in pdf.pages:
        #         text = page.extract_text()
        #         if not text:
        #             continue

        #         lines = [line.strip() for line in text.splitlines() if line.strip()]
        #         capture = False
        #         i = 0

        #         while i < len(lines):
        #             line = lines[i]

        #             # Detect table header start
        #             if not capture and "Package" in line and "Customer" in line and "Part" in line:
        #                 capture = True
        #                 i += 1
        #                 continue

        #             if capture:
        #                 # Match package number or range at line start
        #                 package_match = re.match(r"^(\d+(?:-\d+)?)\s*(.*)$", line)
        #                 if package_match:
        #                     j = i + 1
        #                     package_lines = []

        #                     # Collect lines until next package or end
        #                     while j < len(lines):
        #                         next_line = lines[j]
        #                         # Detect next package start (number or range)
        #                         if re.match(r"^\d+(?:-\d+)?\b", next_line):
        #                             break
        #                         # Skip irrelevant lines
        #                         if re.search(r"TOTAL PACKAGES|PAGE\s*\d+|SHIPPING MARKS|INV\.", next_line, re.IGNORECASE):
        #                             j += 1
        #                             continue
        #                         package_lines.append(next_line)
        #                         j += 1

        #                     # Parse package lines

        #                     current_coo = ""
        #                     # Sometimes COO appears anywhere, so collect all COO lines first
        #                     coo_lines = [pl for pl in package_lines if pl.upper().startswith("MADE IN")]
        #                     if coo_lines:
        #                         current_coo = coo_lines[-1].upper().replace("MADE IN", "").strip()

        #                     # Filter out COO lines from package_lines for parts parsing
        #                     part_lines = [pl for pl in package_lines if not pl.upper().startswith("MADE IN")]

        #                     for pline in part_lines:
        #                         pline = pline.strip()
        #                         if not pline:
        #                             continue

        #                         # Regex to capture parts:
        #                         # Format: PartNumber Description Quantity Weight UnitPrice
        #                         part_pattern = re.match(
        #                             r"^(?P<PartNumber>\S+)\s+(?P<Description>.+?)\s+(?P<Quantity>\d+)\s+(?P<Weight>[\d\.]+)\s+(?P<UnitPrice>[\d\.]+)$",
        #                             pline
        #                         )
        #                         if part_pattern:
        #                             # Extract groups
        #                             part_number = part_pattern.group("PartNumber")
        #                             description = part_pattern.group("Description")
        #                             quantity = part_pattern.group("Quantity")
        #                             weight = part_pattern.group("Weight")
        #                             unit_price = part_pattern.group("UnitPrice")
        #                         else:
        #                             # Fallback split parsing (less reliable but safer)
        #                             parts = pline.split()
        #                             if len(parts) >= 5:
        #                                 part_number = parts[0]
        #                                 quantity = parts[-3]
        #                                 weight = parts[-2]
        #                                 unit_price = parts[-1]
        #                                 description = " ".join(parts[1:-3])
        #                             else:
        #                                 # If line too short, skip to avoid garbage data
        #                                 continue

        #                         package_data.append({
        #                             "PartNumber": part_number,
        #                             "Description": description,
        #                             "Quantity": quantity,
        #                             "Weight": weight,
        #                             "UnitPrice": unit_price,
        #                             "Coo": current_coo
        #                         })
        #                     print("package_data:",package_data)

        #                     i = j
        #                     continue

        #             i += 1


#7

        # package_data = []

        # with pdfplumber.open(saved_pdf_path) as pdf:
        #     for page in pdf.pages:
        #         text = page.extract_text()
        #         if not text:
        #             continue

        #         lines = [line.strip() for line in text.splitlines() if line.strip()]
        #         capture = False
        #         i = 0

        #         while i < len(lines):
        #             line = lines[i]

        #             if not capture and re.search(r"Package\s+Customer\s+Part", line, re.IGNORECASE):
        #                 capture = True
        #                 i += 1
        #                 continue

        #             if capture:
        #                 package_match = re.match(r"^(\d+)\s*(.*)$", line)
        #                 if package_match:
        #                     package_number = package_match.group(1)
        #                     j = i + 1
        #                     package_lines = []

        #                     while j < len(lines):
        #                         next_line = lines[j]
        #                         if re.match(r"^\d+\s*$", next_line):
        #                             break
        #                         if re.search(r"TOTAL PACKAGES|PAGE\s*\d+|SHIPPING MARKS|INV\.", next_line, re.IGNORECASE):
        #                             j += 1
        #                             continue
        #                         package_lines.append(next_line)
        #                         j += 1

        #                     # Initialize variables to avoid UnboundLocalError
        #                     part_number = ""
        #                     description = ""
        #                     quantity = ""
        #                     weight = ""
        #                     unit_price = ""

        #                     customer_part = package_lines[0] if len(package_lines) > 0 else ""
        #                     coo_line = package_lines[1] if len(package_lines) > 1 else ""

        #                     if customer_part.strip():
        #                         part_match = re.match(
        #                             r"(\S+)\s+(.+)\s+(\d+)\s+([\d\.]+)\s+([\d\.]+)",
        #                             customer_part
        #                         )
        #                         if part_match:
        #                             part_number = part_match.group(1)
        #                             description = part_match.group(2)
        #                             quantity = part_match.group(3)
        #                             weight = part_match.group(4)
        #                             unit_price = part_match.group(5)
        #                         else:
        #                             parts = customer_part.split()
        #                             if len(parts) >= 5:
        #                                 part_number = parts[0]
        #                                 description = " ".join(parts[1:-3])
        #                                 quantity = parts[-3]
        #                                 weight = parts[-2]
        #                                 unit_price = parts[-1]

        #                     coo_clean = coo_line.upper().replace("MADE IN", "").strip() if coo_line else ""

        #                     package_data.append({
        #                         "PartNumber": part_number,
        #                         "Description": description,
        #                         "Quantity": quantity,
        #                         "Weight": weight,
        #                         "UnitPrice": unit_price,
        #                         "Coo": coo_clean,
        #                         "PackageNumber": package_number
        #                     })

        #                     i = j
        #                     continue

        #             i += 1

        elif mode == "Nnr_Toshiba_2":
            po_data = []

            with pdfplumber.open(saved_pdf_path) as pdf:
                for page in pdf.pages:
                    text = page.extract_text()
                    if not text:
                        continue
                    lines = [line.strip() for line in text.splitlines() if line.strip()]
                    capture = False
                    i = 0
                    while i < len(lines):
                        line = lines[i]

                        # Detect start of PO table
                        if not capture and re.search(r"Item\s+Customer'?s?\s+P/O\s+No\.?", line, re.IGNORECASE):
                            capture = True
                            i += 1
                            continue

                        if capture:
                            # Match line starting with number + PO number
                            po_match = re.match(r"^\d+\s+([A-Z0-9-]+)\s+(.+)$", line)
                            if po_match:
                                po_number = po_match.group(1)
                                rest = po_match.group(2)

                                # Split from the right to reliably extract Quantity, UnitPrice, Amount
                                try:
                                    description_part, quantity, unit_price, raw_amount = rest.rsplit(maxsplit=3)
                                    description = description_part
                                    amount = re.sub(r"[^\d.,]", "", raw_amount)  # remove asterisks or symbols
                                except ValueError:
                                    # fallback if line structure is irregular
                                    description = rest
                                    quantity = unit_price = amount = "0"

                                # Next line contains Customer PN and Toshiba PN (if exists)
                                customer_pn = ""
                                toshiba_pn = ""
                                if i + 1 < len(lines):
                                    next_line = lines[i + 1].strip()
                                    parts = next_line.split()
                                    if len(parts) >= 2:
                                        customer_pn = parts[0]
                                        toshiba_pn = parts[1].replace("RoHS", "")
                                        i += 1  # skip this line as it's processed
                                    elif len(parts) == 1:
                                        customer_pn = parts[0]
                                        i += 1

                                po_data.append({
                                    "PO": po_number,
                                    "Description": description,
                                    "Quantity": quantity,
                                    "UnitPrice": unit_price,
                                    "Amount": amount,
                                    "CustomerPN": customer_pn,
                                    "ToshibaPN": toshiba_pn
                                })

                        i += 1

            package_data = []
            with pdfplumber.open(saved_pdf_path) as pdf:
                for page in pdf.pages:
                    text = page.extract_text()
                    if not text:
                        continue
                    lines = [line.strip() for line in text.splitlines() if line.strip()]
                    capture = False
                    i = 0
                    while i < len(lines):
                        line = lines[i]
                        # Detect start of package table
            #             if not capture and re.search(r"Package\s+Customer\s+Part", line, re.IGNORECASE):
                        if not capture and "Package" in line and "Customer" in line and "Part" in line:
                            capture = True
                            i += 1
                            continue

                        if capture:
                            # Detect a new package by numeric line
                            #package_match = re.match(r"^(\d+)\s*(.*)$", line)
                            package_match = re.match(r"^(\d+(?:-\d+)?)\s*(.*)$", line)
                            if package_match:
                                package_number = package_match.group(1)
                                j = i + 1
                                package_lines = []


                                # Collect all lines until next package or end
                                while j < len(lines):
                                    next_line = lines[j]
                                    if re.match(r"^\d+\s*$", next_line):  # next package
                                        break
                                    # Skip irrelevant lines
                                    if re.search(r"TOTAL PACKAGES|PAGE\s*\d+", next_line, re.IGNORECASE):
                                        j += 1
                                        continue
                                    package_lines.append(next_line)
                                    # print("package_lines:",package_lines)
                                    j += 1
                                customer_part = package_lines[0] if len(package_lines) > 0 else ""
                                coo = package_lines[1] if len(package_lines) > 1 else ""
                                if customer_part.strip() or coo.strip():
                                    part_pattern = re.match(
                                        r"(\S+)\s+(.+)\s+(\d+)\s+([\d\.]+)\s+([\d\.]+)",customer_part)
                                    if part_pattern:
                                        part_number = part_pattern.group(1)
                                        description = part_pattern.group(2)
                                        quantity = part_pattern.group(3)
                                        weight = part_pattern.group(4)
                                        unit_price = part_pattern.group(5)
                                    else:
                                        parts = customer_part.split()
                                        part_number = parts[0] if len(parts) > 0 else ""
                                        description = " ".join(parts[1:-3]) if len(parts) > 4 else ""
                                        quantity = parts[-3] if len(parts) > 2 else ""
                                        weight = parts[-2] if len(parts) > 1 else ""
                                        unit_price = parts[-1] if len(parts) > 0 else ""
                                    coo_clean = coo.upper().replace("MADE IN ", "").strip()
                                    package_data.append({
                                        "PartNumber":part_number ,
                                        "Description":description,
                                        "Quantity": quantity,
                                        "Weight": weight,
                                        "UnitPrice":unit_price ,
                                        "Coo":coo_clean,

                                    })

                                i = j
                                continue
                        i += 1


            # ===============================
            # AGGREGATE BY CUSTOMERPN
            # ===============================
            aggregated_po_data = {}

            for item in po_data:
                cust_pn = item["CustomerPN"]
                if not cust_pn:
                    continue  # skip empty CustomerPN

                # Safely convert numbers
                try:
                    quantity = float(item["Quantity"].replace(",", ""))
                except:
                    quantity = 0.0
                try:
                    amount = float(item["Amount"].replace(",", ""))
                except:
                    amount = 0.0
                try:
                    unit_price = float(item["UnitPrice"].replace(",", ""))
                except:
                    unit_price = 0.0

                if cust_pn in aggregated_po_data:
                    aggregated_po_data[cust_pn]["Quantity"] += quantity
                    aggregated_po_data[cust_pn]["Amount"] += amount
                    aggregated_po_data[cust_pn]["PO"] += f"/{item['PO']}"
                else:
                    aggregated_po_data[cust_pn] = {
                        "PO": item["PO"],
                        "Description": item["Description"],
                        "Quantity": quantity,
                        "UnitPrice": unit_price,
                        "Amount": amount,
                        "CustomerPN": cust_pn,
                        "ToshibaPN": item["ToshibaPN"]
                    }

            # Convert to list for Excel or further processing
            po_data = list(aggregated_po_data.values())



        # ===============================
        # CREATE EXCEL FILE
        # ===============================

        excel_fs = FileSystemStorage(location='media/excels')
        os.makedirs(excel_fs.location, exist_ok=True)
        excel_filename = os.path.splitext(pdf_file.name)[0] + f"_{mode}_trimmed.xlsx"
        excel_path = os.path.join(excel_fs.location, excel_filename)

        # Create workbook first
        wb = Workbook()

        if mode == "Nnr_Toshiba_1":
            # Sheet 1: PO Data
            ws1 = wb.active
            ws1.title = "PO Data"
            ws1.append(["PO", "Description", "Quantity", "UnitPrice", "Amount", "CustomerPN", "ToshibaPN"])
            for item in po_data:
                ws1.append([
                    item["PO"],
                    item["Description"],
                    item["Quantity"],
                    item["UnitPrice"],
                    item["Amount"],
                    item["CustomerPN"],
                    item["ToshibaPN"]
                ])

            # Sheet 2: Package Data
            ws2 = wb.create_sheet(title="Package Data")
            ws2.append(["PartNumber", "Description", "Quantity", "Weight", "UnitPrice", "Coo"])
            for item in package_data:
                ws2.append([
                    item["PartNumber"],
                    item["Description"],
                    item["Quantity"],
                    item["Weight"],
                    item["UnitPrice"],
                    item["Coo"]
                ])
        elif mode == "Nnr_Toshiba_2":
            # Sheet 1: PO Data
            ws1 = wb.active
            ws1.title = "PO Data"
            ws1.append(["PO", "Description", "Quantity", "UnitPrice", "Amount", "CustomerPN", "ToshibaPN"])
            for item in po_data:
                ws1.append([
                    item["PO"],
                    item["Description"],
                    item["Quantity"],
                    item["UnitPrice"],
                    item["Amount"],
                    item["CustomerPN"],
                    item["ToshibaPN"]
                ])

            # Sheet 2: Package Data
            ws2 = wb.create_sheet(title="Package Data")
            ws2.append(["PartNumber", "Description", "Quantity", "Weight", "UnitPrice", "Coo"])
            for item in package_data:
                ws2.append([
                    item["PartNumber"],
                    item["Description"],
                    item["Quantity"],
                    item["Weight"],
                    item["UnitPrice"],
                    item["Coo"]
                ])

        # Other modes
        else:
            ws = wb.active
            ws.title = f"{mode} Trimmed Data"

            if mode == "Celine":
                ws.append(["COO", "Quantity", "Unit", "Unit Price", "Amount", "HS Code","Invoice Number"])
            elif mode =="Toshiba_Import":
                ws.append(["PO Number","Description","Quantity","Unit","Unit Price","Amount","Diffused In","Assembled In"])
            elif mode == "Nnr_Global_Panasonic":
                ws.append(["COO", "Description", "Quantity", "Unit", "Amount"])
            elif mode == "Marinetrans":
                ws.append(["HSCode", "CoO","Qty","UOM","Weight","W.UOM","Unit Price","Currency","Customs Value"])
            elif mode == "linux":
                ws.append(["HS Code","Description","No. of Items","Unit Value","Total Value","COO"])
            else:
                ws.append(["COO", "Quantity", "Unit", "Unit Price", "Amount"])

            for item in extracted_data:
                if mode == "Celine":
                    ws.append([
                        item["COO"],
                        item["Quantity"],
                        item["Unit"],
                        item["Unit Price"],
                        item["Amount"],
                        item["HS Code"],
                        item["Invoice Number"]
                    ])
                elif mode == "Toshiba_Import":
                    ws.append([
                        item["PO Number"],
                        item["Description"],
                        item["Quantity"],
                        item["Unit"],
                        item["Unit Price"],
                        item["Amount"],
                        item["Diffused In"],
                        item["Assembled In"]
                    ])
                elif mode == "Nnr_Global_Panasonic":
                    ws.append([
                        item["COO"],
                        item["Description"],
                        item["Quantity"],
                        item["Unit"],
                        item["Amount"],
                    ])
                elif mode == "Marinetrans":
                    ws.append([
                        item["HSCode"],
                        item["CoO"],
                        item["Batch Number 2"],
                        item["UOM"],
                        item["Weight"],
                        item["W.UOM"],
                        item["Unit Price"],
                        item["Currency"],
                        item["Customs Value"],
                    ])
                elif mode == "linux":
                    ws.append([
                            item["HS Code"],
                            item["Description"],
                            item["No. of Items"],
                            item["Unit Value"],
                            item["Total Value"],
                            item["COO"]
            ])
                else:
                    ws.append([
                        item["COO"],
                        item["Quantity"],
                        item["Unit"],
                        item["Unit Price"],
                        item["Amount"]
                    ])

        # Save workbook after writing all sheets
        wb.save(excel_path)
        with open(excel_path, "rb") as f:
            response = HttpResponse(
                f.read(),
                content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
            response["Content-Disposition"] = f'attachment; filename="{excel_filename}"'
            return response

    # GET request or no file uploaded
    return render(request, "index.html")











from django.http import JsonResponse

def thosiba_checking(request):
    pdf_path = r"C:\Users\Admin\Downloads\18KA0EL_TEA.pdf"
    extracted_data = []

    with pdfplumber.open(pdf_path) as pdf:
        full_text = ""
        for page in pdf.pages:
            text = page.extract_text()
            if text:
                full_text += text + "\n"

    # Split based on PO numbers
    parts = re.split(r"P/O\s*NO:\s*(\d+)", full_text)

    # Quantity + Unit pattern
    unit_pattern = r"\b(\d{1,3}(?:,\d{3})*|\d+)\s*(PCS|NOS|UNITS|BOX|PACK|SET|PAIR|KG|G|TON|L|ML|M|CM|FT)\b"

    for i in range(1, len(parts), 2):
        po_number = parts[i]
        text_after = parts[i + 1]

        lines = text_after.strip().splitlines()

        description = ""
        quantity = ""
        unit = ""
        unit_price = ""
        amount = ""
        diffused_in = ""
        assembled_in = ""

        for idx, line in enumerate(lines):
            clean_line = line.strip()

            if not clean_line:
                continue

            # ❌ Skip unwanted lines
            if re.search(r"OUR\s+REF\s+NO", clean_line, re.IGNORECASE):
                continue

            if "@" in clean_line:  # skip noisy @ lines
                continue

            # ✅ Description (remove trailing numbers like 0.4 0.008)
            if not description:
                description = re.sub(
                    r"\s+\d+(\.\d+)?(\s+\d+(\.\d+)?)*$",
                    "",
                    clean_line
                )

            # ✅ Quantity + Unit
            qty_match = re.search(unit_pattern, clean_line, re.IGNORECASE)
            if qty_match:
                quantity = qty_match.group(1)   # keep comma format
                unit = qty_match.group(2).upper()

                # ✅ Extract Unit Price & Amount
                after_qty = clean_line[qty_match.end():].strip()

                numbers = re.findall(
                    r"\d{1,3}(?:,\d{3})*\.\d+|\d+\.\d+",
                    after_qty
                )

                if len(numbers) >= 1:
                    unit_price = numbers[0]

                if len(numbers) >= 2:
                    amount = numbers[1]

                # # ✅ Look ahead for Diffused / Assembled (ALL CASES)
                # next_lines = lines[idx+1: idx+6]

                # # Merge lines
                # combined_text = " ".join(
                #     [l.strip() for l in next_lines if l.strip()]
                # )

                # # 🔥 Normalize broken OCR text (A S S E MBLED → ASSEMBLED)
                # combined_text = re.sub(
                #     r'(?<=\b[A-Z])\s+(?=[A-Z]\b)',
                #     '',
                #     combined_text.upper()
                # )

                # # ✅ Extract Diffused
                # diff_match = re.search(
                #     r"DIFFUSED\s+IN:\s*([A-Z]+)",
                #     combined_text
                # )
                # if diff_match:
                #     diffused_in = diff_match.group(1)

                # # ✅ Extract Assembled
                # assm_match = re.search(
                #     r"ASSEMBLED\s+IN:\s*([A-Z]+)",
                #     combined_text
                # )
                # if assm_match:
                #     assembled_in = assm_match.group(1)

                                # ✅ Look ahead lines
                next_lines = lines[idx+1: idx+6]
                combined_text = " ".join([l.strip() for l in next_lines if l.strip()])

                combined_text = combined_text.upper()

                # 🔥 Normalize broken text
                combined_text = re.sub(r'(?<=\b[A-Z])\s+(?=[A-Z]\b)', '', combined_text)

                # ✅ Step 1: Extract Diffused
                diff_match = re.search(r"DIFFUSED\s*IN[:\s]*([A-Z]+)", combined_text)
                if diff_match:
                    diffused_in = diff_match.group(1)

                # ✅ Step 2: Try normal Assembled
                assm_match = re.search(r"ASSEMBLED\s*IN[:\s]*([A-Z]+)", combined_text)

                if assm_match:
                    assembled_in = assm_match.group(1)

                else:
                    # 🔥 Step 3: Fallback → extract all IN values
                    in_values = re.findall(r"IN[:\s]*([A-Z]+)", combined_text)

                    if len(in_values) >= 2:
                        diffused_in = in_values[0]
                        assembled_in = in_values[1]

                break  # stop after correct quantity row

        extracted_data.append({
            "PO Number": po_number,
            "Description": description,
            "Quantity": quantity,
            "Unit": unit,
            "Unit Price": unit_price,
            "Amount": amount,
            "Diffused In": diffused_in,
            "Assembled In": assembled_in
        })

    return JsonResponse({"data": extracted_data})










